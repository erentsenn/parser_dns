import datetime
import pprint
import sqlite3
import types
import asyncio
from yadisk import YaDisk
from telebot.types import Message, InlineKeyboardMarkup, InlineKeyboardButton, ReplyKeyboardMarkup, KeyboardButton, CallbackQuery, ReplyKeyboardRemove
from configparser import ConfigParser
from repository import Users
from telebot.types import LabeledPrice
import json
import os

#calendar
from datetime import date, timedelta
import telebot
from filters import calendar_factory, calendar_zoom, bind_filters
from keyboards import generate_calendar_days, generate_calendar_months, EMTPY_FIELD
from telebot.async_telebot import AsyncTeleBot
from selenium.webdriver.common.by import By
from selenium import webdriver
import pyautogui as pag
import pyperclip
#calendar
import openpyxl
from openpyxl import styles

driver = webdriver.Firefox(executable_path='geckodriver.exe')
provider_token = '381764678:TEST:37080'
con = sqlite3.connect('database.db')
cur = con.cursor()
redFill = styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

y = YaDisk(token='AQAAAAA3RqkuAAfdRhkGvCoYjUYaoDObYD0ehm4')
stek = []

config = ConfigParser()
config.read('config.ini')
admins = list()
for admin in config['data']['admins'].split(','):
    admins.append(int(admin))
token = config['data']['token']
users = Users(con)
stek = list()
bot = AsyncTeleBot(token)
bind_filters(bot)
orders = {}
dates = {}
kill_order = {}
deletes = []
dict_reformat_to_rus = {'Sat': 'сб',
        'Sun': 'вск',
        'Mon': 'пн',
        'Tue': 'вт',
        'Wed': 'ср',
        'Thu': 'чт',
        'Fri': 'пт'}



something = 0
if os.listdir('orders'):
    for path in os.listdir('orders'):
        if int(path.split('.')[0].split('-')[1]) >= something:
            something = int(path.split('.')[0].split('-')[1]) + 1


async def deletes():
    global deletes
    wb = openpyxl.load_workbook(filename='table.xlsx')
    sheet = wb['Суточные и часовые объекты']
    deletes = []
    counter_first = 3
    for row in sheet.iter_rows(min_row=3, max_col=18):
        if row[1].fill.start_color.index == 2 or row[1].fill.start_color.index == 'FF0000':
            deletes.append(counter_first)
        counter_first += 1
    print(deletes)



async def insert_values(stek=[], position='B110'):
    await asyncio.sleep(5)
    pag.moveTo(65, 133)
    pag.click()
    pyperclip.copy(position)
    pag.hotkey('ctrl', 'v')
    pag.press('enter')
    for i in stek:
        await asyncio.sleep(2)
        pyperclip.copy(str(i))
        pag.hotkey('ctrl', 'v')
        pag.press('right')


async def connection():
    driver.fullscreen_window()
    try:
        driver.get('https://disk.yandex.ru/i/U07TEs4L_saexw')
        print('funny')
    except Exception as e:
        print(e)
        driver.save_screenshot('screenshot.png')
    print(driver.service.service_args)


async def update_browser():
    while True:
        await asyncio.sleep(1800)
        pyperclip.copy('A1')
        pag.moveTo(65, 133)
        pag.click()
        pag.hotkey('ctrl', 'v')
        pag.press('enter')
        pyperclip.copy(str(1))
        pag.hotkey('ctrl', 'v')
        pag.press('right')
        print('not_funny')


async def insert_values_globally(stek=[]):
    print(y.download('Статистика аренды ВСЕХ объектов 2022.xlsx', 'table.xlsx'))
    excel_file = openpyxl.load_workbook('table.xlsx')
    excel_sheet = excel_file['Суточные и часовые объекты']
    indexes = []
    t = 3
    for row in excel_sheet.iter_rows(values_only=True, min_row=3, max_col=18):
        k = 0
        reformat_row = list(row)
        for i in range(18):
            if type(reformat_row[i]) == types.NoneType:
                k += 1
        reformat_row.pop(0)
        if k <= 13:
            indexes.append((t, 'not_null'))
        else:
            indexes.append((t, 'null'))
        t += 1
    maxt = 0
    for i in range(len(indexes)):
        if indexes[i][1] == 'not_null':
            maxt = max(maxt, indexes[i][0])
    for row in stek:
        maxt += 1
        print(maxt, row)
        row += [maxt]
        cur.execute(f'''INSERT INTO orders (id, 
                        name, 
                        phone,
                        tariff, 
                        arrival, 
                        eviction, 
                        what_day, 
                        time_arrival, 
                        total_hours, 
                        booking, 
                        debt, 
                        cleaning, 
                        total, 
                        amount, 
                        additionals, 
                        amount_additionals, 
                        comments,
                        xlsx_row) 
                        VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);''',
                    row)
        con.commit()
        await insert_values(stek=row, position=f'B{maxt}')
    pag.hotkey('ctrl', 's')
    await asyncio.sleep(10)
    print(y.download('Статистика аренды ВСЕХ объектов 2022.xlsx', 'table.xlsx'))
    print('uploaded')


async def update_dates():
    print('update_dates')
    for some_order in list(cur.execute("SELECT tariff, arrival, eviction, time_arrival, what_day FROM orders")):
        tarif = some_order[0].lower()
        mode = 'почасово'
        if tarif == 'банчо+':
            tarif = 'банчо +'
        elif '5' in tarif:
            tarif = '5'
        elif '10' in tarif:
            tarif = '10'
        elif '1' in tarif:
            tarif = '1'
        arrival = some_order[1].split()[0]
        year = arrival.split('-')[0]
        month = arrival.split('-')[1]
        day = arrival.split('-')[2]
        arrival_date = date(int(year), int(month), int(day))
        eviction = some_order[2].split()[0]
        if eviction != '0':
            year_ev = eviction.split('-')[0]
            month_ev = eviction.split('-')[1]
            day_ev = eviction.split('-')[2]
            eviction_date = date(int(year_ev), int(month_ev), int(day_ev))
            mode = 'посуточно'

        time = some_order[3]
        if 'сут' in time and eviction == '0':
            eviction_date = arrival_date + timedelta(days=1)
            mode = 'посуточно'

        if mode == 'посуточно':
            try:
                days = (eviction_date - arrival_date).days
                await add_key_to_day(tarif, arrival_date)
                for t in range(14, 24):
                    tt = datetime.time(hour=t, minute=0).strftime('%H.%M')
                    if tt not in dates[tarif][year][month][day]:
                        dates[tarif][year][month][day].append(tt)
                for _ in range(days - 1):
                    arrival_date += timedelta(days=1)
                    await add_key_to_day(tarif, arrival_date)
                    for t in range(0, 24):
                        tt = datetime.time(hour=t, minute=0).strftime('%H.%M')
                        if tt not in dates[tarif][arrival_date.strftime("%Y")][arrival_date.strftime('%m')][arrival_date.strftime("%d")]:
                            dates[tarif][arrival_date.strftime("%Y")][arrival_date.strftime('%m')][arrival_date.strftime("%d")].append(tt)
                arrival_date += timedelta(days=1)
                for t in range(0, 16):
                    tt = datetime.time(hour=t, minute=0).strftime('%H.%M')
                    await add_key_to_day(tarif, arrival_date)
                    if tt not in dates[tarif][arrival_date.strftime("%Y")][arrival_date.strftime('%m')][arrival_date.strftime("%d")]:
                        dates[tarif][arrival_date.strftime("%Y")][arrival_date.strftime('%m')][arrival_date.strftime("%d")].append(tt)
            except Exception as e:
                print(e, some_order)


        elif mode == 'почасово':
            await add_key_to_day(tarif, arrival_date)
            try:
                time_arrival = datetime.datetime(year=int(year),
                                                 month=int(month),
                                                 day=int(day),
                                                 hour=int(time.split('-')[0].split('.')[0]),
                                                 minute=int(time.split('-')[0].split('.')[1]))
                time_eviction = datetime.datetime(year=int(year),
                                                  month=int(month),
                                                  day=int(day),
                                                  hour=int(time.split('-')[1].split('.')[0]),
                                                  minute=int(time.split('-')[1].split('.')[1]))
                if time_eviction.strftime('%H.%M') in ['00.00', '01.00']:
                    time_eviction += timedelta(days=1)
                delta = (time_eviction-time_arrival)
                for i in range(delta.seconds//3600 + 1):
                    await add_key_to_day(tarif, time_arrival)
                    tt = datetime.time(hour=int(time_arrival.strftime('%H')), minute=0).strftime('%H.%M')
                    if tt not in dates[tarif][time_arrival.strftime("%Y")][time_arrival.strftime('%m')][time_arrival.strftime("%d")]:
                        dates[tarif][time_arrival.strftime("%Y")][time_arrival.strftime('%m')][time_arrival.strftime("%d")].append(tt)
                    time_arrival += timedelta(hours=1)
            except Exception as e:
                print(e, some_order)


async def add_key_to_day(tarif, dt):
    year = dt.strftime("%Y")
    month = dt.strftime("%m")
    day = dt.strftime("%d")
    if tarif not in dates.keys():
        dates[tarif] = {}
    if year not in dates[tarif].keys():
        dates[tarif][year] = {}
    if month not in dates[tarif][year].keys():
        dates[tarif][year][month] = {}
    if day not in dates[tarif][year][month].keys():
        dates[tarif][year][month][day] = list()


async def update_database_from_xlsx():
    global deletes
    print('download')
    ids = list()
    for idd in list(cur.execute("SELECT id FROM orders")):
        ids.append(idd[0])
    wb = openpyxl.load_workbook(filename='table.xlsx', data_only=True)
    sheet = wb['Суточные и часовые объекты']
    reformat_data = list()
    row_number = 3
    for row in sheet.iter_rows(values_only=True, min_row=3, max_col=18):
        if row_number not in deletes:
            reformat_row = list(row)
            k = 0
            for i in range(18):
                if type(reformat_row[i]) == types.NoneType:
                    k += 1
                    reformat_row[i] = '0'
                elif type(reformat_row[i]) == datetime.datetime:
                    d = (str(reformat_row[i]))
                    reformat_row[i] = d
            reformat_row.pop(0)
            if len(reformat_row) == 17:
                reformat_row.append(row_number)
            else:
                reformat_row[17] = row_number

            if k <= 13:
                reformat_data.append(reformat_row[:20])
        row_number += 1
    i = 0
    for order in reformat_data:
        i += 1
        if i not in ids:
            zakaz = [i, str(order[1]), str(order[2]), str(order[3]), str(order[4]), str(order[5]), str(order[6]),
                        str(order[7]).replace(':', '.'), str(order[8]), str(order[9]), str(order[10]), str(order[11]), str(order[12]),
                        str(order[13]), str(order[14]), str(order[15]), str(order[16]), order[17]]
            cur.execute(f'''INSERT INTO orders (id, name, phone, tariff, arrival, eviction, what_day, time_arrival, total_hours, booking, debt, cleaning, total, amount, additionals, amount_additionals, comments, xlslx_row) VALUES 
                    (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);''', zakaz)
            con.commit()


# handlers and other shit will be under this comment
@bot.message_handler(commands=['start'])
async def start_message(message: Message):
    await users.insert(message.from_user)
    markup = ReplyKeyboardMarkup()
    markup.add(KeyboardButton('Забронировать'), KeyboardButton('Отменить бронь'))
    markup.add(KeyboardButton('О нас'), KeyboardButton('Позвонить'))
    if message.chat.id in admins:
        markup.add(KeyboardButton('get_analytics'))
        markup.add(KeyboardButton('get_dates'))
    await bot.send_message(message.chat.id, f'Чтобы записаться в bancho, нажмите кнопку "Записаться"',
                        reply_markup=markup)



@bot.message_handler(func=lambda message: message.text == 'Забронировать')
async def choose_tariff(message: Message):
    """0) тариф
    1) список дат
    2) посуточно или почасово
    3) список доп услуг
    4) хз
    5)номер
    6) имя
    7) количество человек
    """
    orders[message.chat.id] = [0] * 8
    orders[message.chat.id][6] = message.from_user.full_name
    markup = ReplyKeyboardMarkup()
    markup.add(KeyboardButton('Посуточно'))
    markup.add(KeyboardButton('Почасово'))

    await bot.send_message(message.chat.id, 'Выберите формат бронирования', reply_markup=markup)


@bot.message_handler(func=lambda message: 'Посуточно' == message.text or 'Почасово' == message.text)
async def choose_type_of_rent(message: Message):
    orders[message.chat.id][2] = message.text.lower()
    a = telebot.types.ReplyKeyboardRemove()
    if message.text == 'Посуточно':
        with open('photos/dacho.jpg', mode='rb') as f:
            markup = InlineKeyboardMarkup()
            markup.add(InlineKeyboardButton('Выбрать тариф ДаЧО', callback_data='tarif дачо'))
            await bot.send_photo(message.chat.id, photo=f, reply_markup=markup)

        with open('photos/5.png', mode='rb') as f:
            markup = InlineKeyboardMarkup()
            markup.add(InlineKeyboardButton('Выбрать эту квартиру', callback_data='tarif 5'))
            markup.add(InlineKeyboardButton(text='ссылка на авито',
                                            url='https://www.avito.ru/elista/kvartiry/1-k._kvartira_24m_44et._2397097954?utm_campaign=native&utm_medium=item_page_ios&utm_source=soc_sharing_seller'))
            await bot.send_photo(message.chat.id, photo=f, reply_markup=markup)

        with open('photos/1.png', mode='rb') as f:
            markup = InlineKeyboardMarkup()
            markup.add(InlineKeyboardButton('Выбрать эту квартиру', callback_data='tarif 1'))
            markup.add(InlineKeyboardButton(text='ссылка на авито',
                                            url='https://www.avito.ru/elista/kvartiry/2-k._kvartira_50m_55et._2428877152?utm_campaign=native&utm_medium=item_page_ios&utm_source=soc_sharing'))
            await bot.send_photo(message.chat.id, photo=f, reply_markup=markup)

        with open('photos/10.png', mode='rb') as f:
            markup = InlineKeyboardMarkup()
            markup.add(InlineKeyboardButton('Выбрать эту квартиру', callback_data='tarif 10'))
            markup.add(InlineKeyboardButton(text='ссылка на авито', url='https://www.avito.ru/elista/kvartiry/3-k._kvartira_75m_89et._2365592849?utm_campaign=native&utm_medium=item_page_ios&utm_source=soc_sharing_seller'))
            await bot.send_photo(message.chat.id, photo=f, reply_markup=markup)

    elif message.text == 'Почасово':
        with open('photos/bancho+.jpg', 'rb') as photo:
            markup = InlineKeyboardMarkup()
            markup.add(InlineKeyboardButton('Выбрать БанЧО +', callback_data='tarif банчо +'))
            await bot.send_photo(message.chat.id, photo, reply_markup=markup)

        with open('photos/bancho.jpg', 'rb') as photo:
            markup = InlineKeyboardMarkup()
            markup.add(InlineKeyboardButton('Выбрать БанЧО', callback_data='tarif банчо'))
            await bot.send_photo(message.chat.id, photo, reply_markup=markup)

        with open('photos/domicho.jpg', 'rb') as photo:
            markup = InlineKeyboardMarkup()
            markup.add(InlineKeyboardButton('Выбрать ДОМиЧО', callback_data='tarif домичо'))
            await bot.send_photo(message.chat.id, photo, reply_markup=markup)



@bot.callback_query_handler(func=lambda call: 'tarif' in call.data)
async def choose_date_pochas(call: CallbackQuery):
    orders[call.message.chat.id][0] = call.data[6:]
    print(orders, call.data)
    now = date.today()
    tarif = call.data[6:]
    print(tarif)
    user_dates = []
    mode = orders[call.message.chat.id][2]
    orders[call.message.chat.id][1] = user_dates
    await bot.send_message(call.message.chat.id, 'Календарь',
                           reply_markup=generate_calendar_days(year=now.year, month=now.month, tarif=tarif, dates=dates,
                                                               user_dates=user_dates, mode=mode))



@bot.callback_query_handler(func=lambda call: call.data == "выбрать время заселения")
async def choose_time_posut(call: CallbackQuery):
    orders[call.message.chat.id][4] = []
    date = orders[call.message.chat.id][1][0]
    if date:
        tarif = orders[call.message.chat.id][0]
        year = date[0]
        month = date[1]
        day = date[2]
        if year not in dates[tarif].keys():
            dates[tarif][year] = {}
        if month not in dates[tarif][year].keys():
            dates[tarif][year][month] = {}
        if day not in dates[tarif][year][month].keys():
            dates[tarif][year][month][day] = list()
        time_markup = InlineKeyboardMarkup()
        dt = datetime.datetime(year=int(year), month=int(month), day=int(day), hour=8)
        button_list = []
        k = 0
        for i in range(9, 26):
            delta = timedelta(hours=1)
            dt += delta
            d = str(dt.day)
            m = str(dt.month)
            y = str(dt.year)
            if len(d) == 1:
                d = '0' + d
            if len(m) == 1:
                m = '0' + m
            h = dt.strftime('%H')

            if y not in dates[tarif].keys():
                dates[tarif][y] = {}
            if m not in dates[tarif][y].keys():
                dates[tarif][y][m] = {}
            if d not in dates[tarif][y][m].keys():
                dates[tarif][y][m][d] = list()
            if f'{h}.00' in dates[orders[call.message.chat.id][0]][y][m][d]:
                button_list.append(InlineKeyboardButton(f'{h}:00' + '❌', callback_data='just nothing'))
            else:
                button_list.append(InlineKeyboardButton(f'{h}:00', callback_data=f'+time {h}.00'))
            k += 1
            if (k) % 3 == 0:
                k = 0
                time_markup.add(*button_list)
                button_list = []
        if button_list:
            time_markup.add(*button_list)
        await bot.send_message(call.message.chat.id, 'выберите время', reply_markup=time_markup)
    else:
        await bot.send_message(call.message.chat.id, 'вы не выбрали дату')

    markup = InlineKeyboardMarkup(row_width=8)
    markup.add(InlineKeyboardButton('Ввести количество человек', callback_data='количество человек'))
    await bot.send_message(call.message.chat.id,
                           'Когда определитесь с выбором, нажмите на кнопку "Ввести количество человек"',
                           reply_markup=markup)


@bot.callback_query_handler(func=lambda call: 'time' in call.data)
async def selecting_time(call: CallbackQuery):
    date = orders[call.message.chat.id][1][0]
    if call.data[0] == '+':
        orders[call.message.chat.id][4].append(call.data.split()[1])
    else:
        orders[call.message.chat.id][4].remove(call.data.split()[1])
    times = orders[call.message.chat.id][4]
    tarif = orders[call.message.chat.id][0]
    year = date[0]
    month = date[1]
    day = date[2]
    print(times)
    if year not in dates[tarif].keys():
        dates[tarif][year] = {}
    if month not in dates[tarif][year].keys():
        dates[tarif][year][month] = {}
    if day not in dates[tarif][year][month].keys():
        dates[tarif][year][month][day] = list()
    time_markup = InlineKeyboardMarkup()
    dt = datetime.datetime(year=int(year), month=int(month), day=int(day), hour=8)
    button_list = []
    k = 0
    for i in range(9, 26):
        k += 1
        delta = timedelta(hours=1)
        dt += delta
        d = str(dt.day)
        m = str(dt.month)
        y = str(dt.year)
        h = dt.strftime('%H')
        if len(d) == 1:
            d = '0' + d
        if len(m) == 1:
            m = '0' + m
        if y not in dates[tarif].keys():
            dates[tarif][y] = {}
        if m not in dates[tarif][y].keys():
            dates[tarif][y][m] = {}
        if d not in dates[tarif][y][m].keys():
            dates[tarif][y][m][d] = list()
        if f'{h}.00' in dates[orders[call.message.chat.id][0]][y][m][d]:
            button_list.append(InlineKeyboardButton(f'{h}:00' + '❌', callback_data='just nothing'))
        elif f'{h}.00' in times:
            button_list.append(InlineKeyboardButton(f'{h}:00' + '✅', callback_data=f'-time {h}.00'))
        else:
            button_list.append(InlineKeyboardButton(f'{h}:00', callback_data=f'+time {h}.00'))
        if (k) % 3 == 0:
            k = 0
            time_markup.add(*button_list)
            button_list = []
    if button_list:
        time_markup.add(*button_list)
    await bot.edit_message_reply_markup(call.message.chat.id, call.message.id,
                                        reply_markup=time_markup)



@bot.callback_query_handler(func=lambda call: call.data=='just nothing' or call.data == 'nothing')
async def calendar_action_handler(call: CallbackQuery):
    await bot.answer_callback_query(call.id)


@bot.callback_query_handler(func=None, calendar_config=calendar_factory.filter())
async def calendar_action_handler(call: telebot.types.CallbackQuery):
    callback_data: dict = calendar_factory.parse(callback_data=call.data)
    tarif = orders[call.message.chat.id][0]
    year, month = int(callback_data['year']), int(callback_data['month'])
    user_dates = orders[call.message.chat.id][1]
    await bot.edit_message_reply_markup(call.message.chat.id, call.message.id,
                                        reply_markup=generate_calendar_days(year=year, month=month, tarif=tarif,
                                                                            dates=dates, user_dates=user_dates,
                                                                            mode=orders[call.message.chat.id][2]))


@bot.callback_query_handler(func=None, calendar_zoom_config=calendar_zoom.filter())
async def calendar_zoom_out_handler(call: telebot.types.CallbackQuery):
    callback_data: dict = calendar_zoom.parse(callback_data=call.data)
    year = int(callback_data.get('year'))
    await bot.edit_message_reply_markup(call.message.chat.id, call.message.id,
                                        reply_markup=generate_calendar_months(year=year))


@bot.callback_query_handler(func=lambda call: call.data == EMTPY_FIELD)
async def callback_empty_field_handler(call: telebot.types.CallbackQuery):
    await bot.answer_callback_query(call.id)


@bot.callback_query_handler(func=lambda call: 'selected date' in call.data)
async def dates_selected(call: telebot.types.CallbackQuery):
    await bot.answer_callback_query(call.id)
    tarif = orders[call.message.chat.id][0]
    year = call.data.split()[-3]
    month = call.data.split()[-2]
    day = call.data.split()[-1]
    if orders[call.message.chat.id][2] == 'почасово':
        orders[call.message.chat.id][1] = []
        orders[call.message.chat.id][4] = []
    elif orders[call.message.chat.id][2] == 'посуточно':
        if month not in dates[tarif][year].keys():
            dates[tarif][year][month] = {}
    orders[call.message.chat.id][1].append((year, month, day))
    user_dates = orders[call.message.chat.id][1]
    print(user_dates)
    now = datetime.date.today()
    await bot.edit_message_reply_markup(chat_id=call.message.chat.id,
                                        reply_markup=generate_calendar_days(year=now.year, month=int(month),
                                                                            tarif=tarif, dates=dates,
                                                                            user_dates=user_dates,
                                                                            mode=orders[call.message.chat.id][2]),
                                        message_id=call.message.id)
    if orders[call.message.chat.id][2] == 'почасово':
        markup = InlineKeyboardMarkup(row_width=8)
        markup.add(InlineKeyboardButton("Выбрать время заселения", callback_data='выбрать время заселения'))
        curr_m = datetime.date(year=int(year), month=int(month), day=int(day)).strftime("%d %B")
        await bot.send_message(call.message.chat.id, f'Выбрана дата: {curr_m}, нажмите на кнопку "Выбрать время заселения", чтобы выбрать интересующее вас время на текущую дату',
                               reply_markup=markup)
    elif orders[call.message.chat.id][2] != 'почасово' and len(user_dates) == 1:
        markup = InlineKeyboardMarkup(row_width=8)
        markup.add(InlineKeyboardButton('Ввести количество человек', callback_data='количество человек'))
        await bot.send_message(call.message.chat.id,
                               'Когда определитесь с выбором, нажмите на кнопку "Ввести количество человек"',
                               reply_markup=markup)


@bot.callback_query_handler(func=lambda call: 'unselect date' in call.data)
async def unselecting_date(call: telebot.types.CallbackQuery):
    await bot.answer_callback_query(call.id)
    year = call.data.split()[-3]
    month = call.data.split()[-2]
    day = call.data.split()[-1]
    tarif = orders[call.message.chat.id][0]
    orders[call.message.chat.id][1].remove((year, month, day))
    user_dates = orders[call.message.chat.id][1]
    print(user_dates)
    now = date.today()
    await bot.edit_message_reply_markup(chat_id=call.message.chat.id,
                                        reply_markup=generate_calendar_days(year=now.year, month=int(month), tarif=tarif,
                                                                            dates=dates, user_dates=user_dates,
                                                                            mode=orders[call.message.chat.id][2]),
                                        message_id=call.message.id)

# calendar handlers
@bot.callback_query_handler(func=lambda call: call.data == 'количество человек')
async def select_people_count(call: CallbackQuery):
    markup = InlineKeyboardMarkup()
    markup.add(InlineKeyboardButton('+1', callback_data='+1 mans'))
    markup.add(InlineKeyboardButton(f'{orders[call.message.chat.id][7]}', callback_data='just_nothing'))
    markup.add(InlineKeyboardButton('-1', callback_data='-1 mans'))
    await bot.send_message(call.message.chat.id, 'Введите желаемое количество человек', reply_markup=markup)
    if orders[call.message.chat.id][0] not in ['10', '5', '1']:
        markup = InlineKeyboardMarkup()
        markup.add(InlineKeyboardButton('Выбрать дополнительные услуги', callback_data='выбрать допы'))
        await bot.send_message(call.message.chat.id,
                               'Когда определитесь с выбором, нажмите на кнопку "Выбрать дополнительные услуги"',
                               reply_markup=markup)
    else:
        m = InlineKeyboardMarkup()
        m.add(InlineKeyboardButton('Завершить заказ', callback_data='Завершить заказ'))
        await bot.send_message(call.message.chat.id,
                               'Когда определитесь с выбором, нажмите на кнопку "Завершить заказ"',
                               reply_markup=m)




@bot.callback_query_handler(func=lambda call: 'mans' in call.data)
async def add_or_minus_people(call: CallbackQuery):
    max_count_mans = {'10': (7, 7, 0),
                      'банчо': (7, 20, 300),
                      'банчо +': (8, 20, 300),
                      'дачо': (100, 100, 0),
                      'домичо': (8, 15, 200),
                      '5': (7, 7, 0),
                      '1': (4, 4, 0)
                      }
    if '+1' in call.data:
        tarif = orders[call.message.chat.id][0]
        orders[call.message.chat.id][7] += 1
        markup = InlineKeyboardMarkup()
        mans = orders[call.message.chat.id][7]

        if max_count_mans[tarif][1] == mans:
            await bot.send_message(call.message.chat.id, 'Достигнут лимит количества человек!')
        elif max_count_mans[tarif][0] <= mans:
            if max_count_mans[tarif][2]:
                plusmans = mans - max_count_mans[tarif][0] + 1
                time = orders[call.message.chat.id][4]
                await bot.send_message(call.message.chat.id, f'за {plusmans} человека нужно будет доплатить {max_count_mans[tarif][2] * plusmans * len(time)}')
            markup.add(InlineKeyboardButton('+1', callback_data='+1 mans'))
        else:
            markup.add(InlineKeyboardButton('+1', callback_data='+1 mans'))
        markup.add(InlineKeyboardButton(f'{orders[call.message.chat.id][7]}', callback_data='just_nothing'))
        markup.add(InlineKeyboardButton('-1', callback_data='-1 mans'))
        try:
            await bot.edit_message_reply_markup(chat_id=call.message.chat.id,
                                                reply_markup=markup,
                                                message_id=call.message.id)
        except Exception as e:
            print('pass')
    elif '-1' in call.data:
        orders[call.message.chat.id][7] -= 1
        mans = orders[call.message.chat.id][7]
        markup = InlineKeyboardMarkup()
        markup.add(InlineKeyboardButton('+1', callback_data='+1 mans'))
        markup.add(InlineKeyboardButton(f'{orders[call.message.chat.id][7]}', callback_data='just_nothing'))
        if mans == 0:
            await bot.send_message(call.message.chat.id, 'Достигнут лимит количества человек!')
        else:
            markup.add(InlineKeyboardButton('-1', callback_data='-1 mans'))
        try:
            await bot.edit_message_reply_markup(chat_id=call.message.chat.id,
                                                reply_markup=markup,
                                                message_id=call.message.id)
        except Exception as e:
            print('pass')
    await bot.answer_callback_query(call.id)




@bot.callback_query_handler(func=lambda call: call.data == 'выбрать допы')
async def select_additionals(call: CallbackQuery):
    orders[call.message.chat.id][3] = []
    markup = InlineKeyboardMarkup()

    markup.add(InlineKeyboardButton('Биокамин - 1000 рублей', callback_data='+ads - Биокамин'))
    markup.add(InlineKeyboardButton('Кальян - 1000 рублей', callback_data='+ads - Кальян'))
    markup.add(InlineKeyboardButton('Веники запаренные - 500 рублей', callback_data='+ads - Веники запаренные'))
    if orders[call.message.chat.id][0] == 'БанЧО +':
        markup.add(InlineKeyboardButton('Свои веники - бесплатно', callback_data='+ads - Свои веники'))
    else:
        markup.add(InlineKeyboardButton('Свои веники - 1000 рублей', callback_data='+ads - Свои веники'))
    markup.add(InlineKeyboardButton('Масло для бани - 200 рублей', callback_data='+ads - Масло для бани'))
    markup.add(InlineKeyboardButton('Простыни, одноразовые тапочки - 200 рублей',
                                    callback_data='+ads - Простыни, одноразовые тапочки'))
    markup.add(InlineKeyboardButton('Пивной диспенсор - 500 рублей', callback_data='+ads - Пивной диспенсор'))
    markup.add(InlineKeyboardButton('Халат - 300 рублей', callback_data='+ads - Халат'))
    markup.add(InlineKeyboardButton('Костровая чаша - 1000 рублей', callback_data='+ads - Костровая чаша'))
    markup.add(InlineKeyboardButton('Массажер для ног - 1000 рублей/час', callback_data='+ads - Массажер для ног'))
    with open('photos/additionals.jpg', 'rb') as photo:
        await bot.send_photo(call.message.chat.id, photo)
    await bot.send_message(call.message.chat.id, 'Выберите дополнительные услуги:', reply_markup=markup)
    m = InlineKeyboardMarkup()
    m.add(InlineKeyboardButton('Завершить заказ', callback_data='Завершить заказ'))
    await bot.send_message(call.message.chat.id, 'Когда определитесь с выбором, нажмите на кнопку "Завершить заказ"',
                           reply_markup=m)


@bot.callback_query_handler(lambda call: 'ads' in call.data)
async def additionals_selecting(call: CallbackQuery):
    tarif = orders[call.message.chat.id][0]
    if call.data[0] == '+':
        orders[call.message.chat.id][3].append(call.data.split('-')[1].lstrip())
    else:
        try:
            orders[call.message.chat.id][3].remove(call.data.split('-')[2].lstrip())
        except Exception as e:
            print(e)
    markup = InlineKeyboardMarkup()
    for i in ['Биокамин - 1000', 'Кальян - 1000', 'Веники запаренные - 500', 'Свои веники', 'Масло для бани - 200',
              'Простыни, одноразовые тапочки - 200', 'Пивной диспенсор - 500', 'Халат - 300',
              'Костровая чаша - 1000 рублей', 'Массажер для ног - 1000 рублей/час']:
        additional = i.split('-')[0].rstrip()
        if tarif == 'банчо +' and i == 'Свои веники':
            i += '- бесплатно'
        elif tarif != 'банчо +' and i == 'Свои веники':
            i += '- 1000 рублей'
        if additional not in orders[call.message.chat.id][3]:
            markup.add(InlineKeyboardButton(i, callback_data=f'+ads - {additional}'))
        else:
            markup.add(InlineKeyboardButton(i+'✅', callback_data=f'-ads - {additional}'))
    try:
        await bot.edit_message_reply_markup(chat_id=call.message.chat.id,
                                            reply_markup=markup,
                                            message_id=call.message.id)
    except Exception as e:
        await bot.answer_callback_query(call.id)


@bot.callback_query_handler(func=lambda call: call.data == 'Завершить заказ')
async def sending_order_to_admin(call: CallbackQuery):
    if orders[call.message.chat.id][0] and orders[call.message.chat.id][1]:
        await bot.send_message(call.message.chat.id, 'Для завершения заказа отправьте ваш номер телефона в формате 8хххххххххх')
    else:
        await bot.send_message(call.message.chat.id,
                               'Вы не выбрали тариф, попробуйте забронировать снова')



@bot.message_handler(func=lambda message: len(message.text) == 11 and message.text[0] =='8')
async def really_finish(message: Message):
    if message.chat.id in orders.keys():
        orders[message.chat.id][5] = message.text
        markup = ReplyKeyboardMarkup()
        markup.add(KeyboardButton('Забронировать'), KeyboardButton('Отменить бронь'))
        markup.add(KeyboardButton('О нас'), KeyboardButton('Позвонить'))
        if message.chat.id in admins:
            markup.add(KeyboardButton('get_analytics'))
            markup.add(KeyboardButton('get_dates'))
        await bot.send_message(message.chat.id, 'Ваша заявка на рассмотрении, вам придет сообщение об оплате, или с вами свяжется администратор', reply_markup=markup)
        tarif = orders[message.chat.id][0]
        additionals = orders[message.chat.id][3]
        year = min([int(i[0]) for i in orders[message.chat.id][1]])
        month = min([int(i[1]) for i in orders[message.chat.id][1] if int(i[0]) == year])
        booking = 0
        cleaning = 0
        ads = ''
        amount_ads = 0
        dict_reformat_additionals_to_ads = {'Биокамин':['бк', 1000],
                                            'Кальян':['к', 1000],
                                            'Веники запаренные':['в', 500],
                                            'Масло для бани':['мс', 200],
                                            'Простыни, одноразовые тапочки':['п', 200],
                                            'Пивной диспенсор':['д', 500],
                                            'Халат':['х', 500],
                                            'Свои веники':['Свои веники', 1000],
                                            'Костровая чаша':['чаша', 1000],
                                            'Массажер для ног': ['мж', 1000],
                                            'Плов - полный казан':['плов полный', 5000],
                                            'Плов - половина казана': ['плов половина', 3000]
                                            }
        if additionals:
            for a in additionals:
                ads += dict_reformat_additionals_to_ads[a][0]
                ads += ' '
                if a == 'Свои веники' and tarif != 'банчо +':
                    amount_ads += 1000
                else:
                    amount_ads += dict_reformat_additionals_to_ads[a][1]
        if tarif in ['дачо', 'домичо', '1']:
            booking = 2000
        elif tarif in ['10', '5']:
            booking = 3000
        elif tarif not in ['10', '5', '1']:
            cleaning = 300
        await bot.send_message(message.chat.id, f'бронирование: {booking}, уборка: {cleaning}, доп.услуги:{amount_ads}')
        with open(f'orders/{message.chat.id}-{something}.json', 'w') as outfile:
            json.dump([something] + orders[message.chat.id], outfile)
        dates = orders[message.chat.id][1]
        mode_dates = orders[message.chat.id][2]
        adds = orders[message.chat.id][3]
        if orders[message.chat.id][4]:
            time = orders[message.chat.id][4]
        else:
            time = 'сутки'
        text = f'поступила бронь на {tarif} {mode_dates}, даты - {dates}, время - {time}, доп.услуги - {adds}, телефон - {message.text}'
        markup = InlineKeyboardMarkup()
        markup.add(InlineKeyboardButton('одобрить', callback_data=f'good : {message.chat.id}-{something}'))
        markup.add(InlineKeyboardButton('отклонить', callback_data=f'bad : {message.chat.id}-{something}'))
        for admin in admins:
            await bot.send_message(admin, text, reply_markup=markup)
        orders.pop(message.chat.id)
    else:
        try:
            sql = f"""SELECT tariff, arrival, eviction, time_arrival, xlslx_row FROM orders WHERE phone == {message.text}"""
            for row in list(cur.execute(sql)):
                text = ''
                dat = row[1].split()[0].split('-')
                m = dat[1]
                d = dat[2]
                if row[2] != '0':
                    edat = row[2].split()[0].split('-')
                    mm = edat[1]
                    dd = edat[2]
                    text = f'''тариф - {row[0]}, даты - с {d}.{m} до {dd}.{mm}'''
                else:
                    text = f'''тариф - {row[0]}, дата - {d}.{m}, на время: {row[3]}'''
                if datetime.date(year=int(dat[0]), month=int(m), day=int(d)) >= datetime.date.today():
                    markup = InlineKeyboardMarkup()
                    markup.add(InlineKeyboardButton(text='Снять бронь', callback_data=f'del_order {row[4]}'))
                    await bot.send_message(message.chat.id, text, reply_markup=markup)
        except Exception as e:
            print(e)
            await bot.send_message(message.chat.id, 'Неправильно введен номер, попробуйте еще раз или свяжитесь с техподдержкой')
        pass



@bot.pre_checkout_query_handler(func=lambda query: True)
async def checkout(pre_checkout_query):
    print('wow')
    await bot.answer_pre_checkout_query(pre_checkout_query.id, ok=True,
                                  error_message="Aliens tried to steal your card's CVV, but we successfully protected your credentials,"
                                                " try to pay again in a few minutes, we need a small rest.")


@bot.message_handler(content_types=['successful_payment'])
async def got_payment(message: Message):
    await bot.send_message(message.chat.id,
                     'Thanks for payment! ')
    for admin in admins:
        await bot.send_message(admin, f'{message.from_user.full_name} оплатил бронирование на {orders[message.chat.id][0]}')



@bot.callback_query_handler(func=lambda call: ('good' in call.data or 'bad' in call.data) and call.from_user.id in admins)
async def thinking(call: CallbackQuery):
    print(call.data)
    file = call.data.split(':')[1].lstrip()
    if 'good' in call.data:
        if os.path.exists(f'orders/{file + "-zapisano"}.json'):
            answ = ''
            await bot.send_message(call.message.chat.id, 'заказ уже записан в яндекс диск')
        else:
            await bot.delete_message(chat_id=call.message.chat.id, message_id=call.message.id)
            await bot.send_message(call.message.chat.id, 'заказ будет записан в яндекс диск')
            answ = 'принято'
    else:
        answ = 'отклонено'
    if answ == 'принято':
        id = int(file.split('-')[0])
        await bot.send_message(id, 'Заявка принята')
        with open(f'orders/{file}.json') as jfile:
            order = json.load(jfile)
        os.rename(f'orders/{file}.json', f'orders/{file + "zapisano"}.json')
        name = order[7]
        phone = order[6]
        tarif = order[1]
        time = order[5]
        additionals = order[4]
        mans = order[8]
        year = min([int(i[0]) for i in order[2]])
        month = min([int(i[1]) for i in order[2] if int(i[0]) == year])
        day = min([int(i[2]) for i in order[2] if int(i[0]) == year and int(i[1]) == month])
        print(order)
        if len(order[2]) != 1:
            year_ev = max([int(i[0]) for i in order[2]])
            month_ev = max([int(i[1]) for i in order[2] if int(i[0]) == year_ev])
            day_ev = max([int(i[2]) for i in order[2] if int(i[0]) == year_ev and int(i[1]) == month_ev])
            dt2 = date(year=int(year_ev), month=int(month_ev), day=int(day_ev))
            if len(str(month_ev)) == 1:
                month_evd = '0' + str(month_ev)
            else:
                month_evd = str(month_ev)
            if len(str(day_ev)) == 1:
                day_evd = '0' + str(day_ev)
            else:
                day_evd = str(day_ev)
            s_ev = f'{year_ev}-{month_evd}-{day_evd} 00:00:00'
            what_day1 = dict_reformat_to_rus[date(year=int(year), month=int(month), day=int(day)).strftime('%a')]
            what_day2 = dict_reformat_to_rus[dt2.strftime('%a')]
            what_day = f'{what_day1}-{what_day2}'
        else:
            s_ev = '0'
            what_day = '0'

        if time:
            if '00:00' in time:
                time.remove('00:00')
                time.append('24:00')
            if '01:00' in time:
                time.remove('01:00')
                time.append('25:00')
            time = sorted(time, key=lambda x: int(x.split('.')[0]))
            maxt = time[-1]
            mint = time[0]
            if maxt == '24.00':
                maxt = '00.00'
            elif maxt == '25.00':
                maxt = '01.00'
            if mint == '24.00':
                mint = '00.00'
            elif mint == '25.00':
                mint = '01.00'
            time_arrival = f'{mint}-{maxt}'
        else:
            time_arrival = 'сутки'
        booking = 0
        debt = 0
        cleaning = 0
        total = 0
        ads = ''
        amount_ads = 0
        comments = ''
        dict_reformat_additionals_to_ads = {'Биокамин':['бк', 1000],
                                            'Кальян':['к', 1000],
                                            'Веники запаренные':['в', 500],
                                            'Масло для бани':['мс', 200],
                                            'Простыни, одноразовые тапочки':['п', 200],
                                            'Пивной диспенсор':['д', 500],
                                            'Халат':['х', 500],
                                            'Свои веники':['Свои веники', 1000],
                                            'Костровая чаша':['чаша', 1000],
                                            'Массажер для ног': ['мж', 1000],
                                            'Плов - полный казан':['плов полный', 5000],
                                            'Плов - половина казана': ['плов половинав', 3000]
                                            }
        if additionals:
            for a in additionals:
                ads += dict_reformat_additionals_to_ads[a][0]
                ads += ' '
                if a == 'Свои веники' and tarif != 'банчо +':
                    amount_ads += 1000
                else:
                    amount_ads += dict_reformat_additionals_to_ads[a][1]
        if tarif in ['дачо', 'домичо']:
            booking = 2000
        elif tarif in ['10', '5', '1']:
            booking = 3000
        if tarif == 'банчо +':
            total += 3000 * len(time)
            if mans >= 8:
                total += 300 * len(time) * (mans - 8)
        elif tarif == 'дачо':
            year_ev = max([int(i[0]) for i in order[2]])
            month_ev = max([int(i[1]) for i in order[2] if int(i[0]) == year_ev])
            day_ev = max([int(i[2]) for i in order[2] if int(i[0]) == year_ev and int(i[1]) == month_ev])
            dt1 = date(year=int(year), month=int(month), day=int(day))
            dt2 = date(year=int(year_ev), month=int(month_ev), day=int(day_ev))
            delta = dt2 - dt1
            deltadays = delta.days
            if dt1 == dt2:
                deltadays = 1
            for i in range(deltadays):
                d = timedelta(days=1)
                dt1 += d
                wd = dict_reformat_to_rus[date(year=int(year), month=int(month), day=int(day)).strftime('%a')]
                if wd in ['пн', 'вт', 'ср', 'чт', 'сб']:
                    total += 7000
                else:
                    total += 9000

        elif tarif == '10':
            year_ev = max([int(i[0]) for i in order[2]])
            month_ev = max([int(i[1]) for i in order[2] if int(i[0]) == year_ev])
            day_ev = max([int(i[2]) for i in order[2] if int(i[0]) == year_ev and int(i[1]) == month_ev])
            dt1 = date(year=int(year), month=int(month), day=int(day))
            dt2 = date(year=int(year_ev), month=int(month_ev), day=int(day_ev))
            delta = dt2 - dt1
            deltadays = delta.days
            if dt1 == dt2:
                deltadays = 1
            for i in range(deltadays):
                total += 4000
        elif tarif == '5':
            year_ev = max([int(i[0]) for i in order[2]])
            month_ev = max([int(i[1]) for i in order[2] if int(i[0]) == year_ev])
            day_ev = max([int(i[2]) for i in order[2] if int(i[0]) == year_ev and int(i[1]) == month_ev])
            dt1 = date(year=int(year), month=int(month), day=int(day))
            dt2 = date(year=int(year_ev), month=int(month_ev), day=int(day_ev))
            delta = dt2 - dt1
            deltadays = delta.days
            if dt1 == dt2:
                deltadays = 1
            for i in range(deltadays):
                total += 4000
        elif tarif == '1':
            total = 4000 * len(order[2])


        elif tarif == 'домичо':
            total += 2000 * len(time)
            if mans >= 15:
                total += 200 * len(time) * (mans - 15)
        elif tarif == 'банчо':
            total += 2000 * len(time)
            if mans >= 5:
                total += 300 * len(time) * (mans - 5)

        elif tarif == 'банчо':
            total += 2000 * len(time)
            if mans >= 5:
                total += 300 * len(time) * (mans - 5)
        if tarif not in ['10', '5']:
            cleaning = 300
        debt = total + cleaning - amount_ads - booking
        prices = [LabeledPrice(label='бронирование', amount=int(booking) * 100),
                  LabeledPrice('уборка', int(cleaning) * 100)]
        await bot.send_invoice(
            chat_id=id,  # chat_id
            title=f'Бронирование {tarif}',  # title
            description=f'some_description',  # description
            invoice_payload='booking',
            provider_token=provider_token,  # provider_token
            currency='rub',  # currency
            prices=prices,  # prices
            is_flexible=False,  # True If you need to set up Shipping Fee
            start_parameter='time-machine-example')

        a = cur.execute('SELECT * FROM orders ORDER BY id DESC LIMIT 1;')
        idd = 100
        for i in a:
            idd = int(i[0]) + 1
        if len(str(month)) == 1:
            month = '0' + str(month)
        else:
            month = str(month)
        if len(str(day)) == 1:
            day = '0' + str(day)
        else:
            day = str(day)
        zakaz = [idd, name, phone, tarif, f'{year}-{month}-{day} 00:00:00', s_ev, what_day, time_arrival,
                 len(str(time)), booking, debt, cleaning, total, mans, ads, amount_ads, comments]
        await insert_values_globally(stek=[zakaz])
        await update_database_from_xlsx()
        await update_dates()
    else:
        id = int(file.split('-')[0])
        await bot.send_message(id, 'Заявка отклонена')


@bot.message_handler(func=lambda message: message.text =='Позвонить нам')
async def call_us(message: Message):
    await bot.send_message(message.chat.id, 'some_phones_and_contacts')


@bot.message_handler(func=lambda message: message.text =='О нас')
async def about(message: Message):
    await bot.send_message(message.chat.id, 'some_inf_about')


@bot.message_handler(func=lambda message: message.text =='Отменить бронь')
async def delete_br(message: Message):
    await bot.send_message(message.chat.id, 'введите номер телефона, на который вы забронировали, в формате 8xxxxxxxxxx')


@bot.callback_query_handler(func=lambda call: 'del_order' in call.data)
async def del_order(call: CallbackQuery):
    kill_order[call.message.chat.id] = call.data.split()[1]
    print(kill_order)
    markup = ReplyKeyboardMarkup()
    markup.add(KeyboardButton('Да, я действительно хочу продолжить'))
    markup.add(KeyboardButton('Нет, я не хочу снимать бронь'))
    await bot.send_message(call.message.chat.id, 'Вы действительно хотите отменить бронирование?', reply_markup=markup)


@bot.message_handler(func=lambda message: message.text == 'Да, я действительно хочу продолжить')
async def ya_uzhe_ne_znau_kak_nazvat(message: Message):
    await bot.send_message(message.chat.id, 'Бронирование успешно снято!')
    markup = ReplyKeyboardMarkup()
    markup.add(KeyboardButton('Забронировать'), KeyboardButton('Отменить бронь'))
    markup.add(KeyboardButton('О нас'), KeyboardButton('Позвонить'))
    if message.chat.id in admins:
        markup.add(KeyboardButton('get_analytics'))
        markup.add(KeyboardButton('get_dates'))
    await bot.send_message(message.chat.id, f'Чтобы записаться в bancho, нажмите кнопку "Записаться"',
                           reply_markup=markup)
    await remove_order_from_cloud(kill_order[message.chat.id])



@bot.message_handler(func=lambda message: message.text == 'Нет, я не хочу снимать бронь')
async def refusing_otmena_broni(message: Message):
    kill_order.pop(message.chat.id)
    markup = ReplyKeyboardMarkup()
    markup.add(KeyboardButton('Забронировать'), KeyboardButton('Отменить бронь'))
    markup.add(KeyboardButton('О нас'), KeyboardButton('Позвонить'))
    if message.chat.id in admins:
        markup.add(KeyboardButton('get_analytics'))
        markup.add(KeyboardButton('get_dates'))
    await bot.send_message(message.chat.id, f'Чтобы записаться в bancho, нажмите кнопку "Записаться"',
                           reply_markup=markup)


async def remove_order_from_cloud(row_number):
    try:
        await remove_order_local(row_number)
        pag.moveTo(231, 102)
        pag.click()
        await asyncio.sleep(1)
        pag.moveTo(251, 305)
        pag.click()
        await asyncio.sleep(1)
        pag.moveTo(65, 133)
        pag.click()
        pyperclip.copy(f'B{row_number}')
        pag.hotkey('ctrl', 'v')
        pag.press('enter')
        pag.moveTo(213, 105)
        for i in range(18):
            pag.click()
            pag.press('right')
            await asyncio.sleep(1)
    except Exception as e:
        print(e)


async def remove_order_local(row_number):
    try:
        sql_delete_query = f"""DELETE FROM orders WHERE xlslx_row = {row_number}"""
        wb = openpyxl.load_workbook(filename='table.xlsx', data_only=True)
        sheet = wb['Суточные и часовые объекты']
        for cell in sheet[int(row_number)]:
            cell.fill = redFill
        wb.save('table.xlsx')
        cur.execute(sql_delete_query)
        con.commit()
    except Exception as e:
        print(e)


async def pre_podgotovka_dannix():
    y.download('Статистика аренды ВСЕХ объектов 2022.xlsx', 'table.xlsx')
    await deletes()
    await update_database_from_xlsx()
    await update_dates()


async def main():
    task1 = asyncio.create_task(
        connection())

    task2 = asyncio.create_task(
        update_browser())

    task3 = asyncio.create_task(
        bot.infinity_polling())

    task4 = asyncio.create_task(
        pre_podgotovka_dannix()
    )

    await task4
    await task1
    await task2
    await task3



asyncio.run(main())